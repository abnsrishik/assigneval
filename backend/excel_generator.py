"""
Excel Generator Module - Assignment Evaluator
=============================================
Generates formatted Excel result sheets using openpyxl.

The output Excel file includes:
- Assignment summary at the top
- Student results table with color-coded grades
- Statistics: average, highest, lowest, pass/fail count
- Individual feedback for each student
"""

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os
import tempfile
from datetime import datetime


def generate_excel(assignment, submissions):
    """
    Generate a formatted Excel results sheet.
    
    Args:
        assignment (dict): Assignment details from database
        submissions (list): List of submission dicts from database
    
    Returns: filepath to the generated .xlsx file
    """
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"
    
    # Define color palette
    NAVY      = "0D1B2A"
    BLUE      = "1565C0"
    ORANGE    = "FF6B35"
    WHITE     = "FFFFFF"
    LIGHT_BG  = "F0F4F8"
    GREEN_BG  = "E8F5E9"
    RED_BG    = "FFEBEE"
    YELLOW_BG = "FFFDE7"
    GRAY      = "90A4AE"

    # ── Header Section ────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = "ASSIGNMENT EVALUATOR — RESULTS"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:I2")
    ws["A2"] = f"SRM Institute of Science and Technology, Ramapuram"
    ws["A2"].font = Font(name="Calibri", size=11, color=GRAY)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Assignment details
    details = [
        ("Assignment:", assignment.get("title", "N/A")),
        ("Subject:", assignment.get("subject", "N/A")),
        ("Teacher:", assignment.get("teacher_name", "N/A")),
        ("Max Marks:", str(assignment.get("max_marks", 100))),
        ("Deadline:", assignment.get("deadline", "N/A")[:16].replace("T", " ")),
        ("Generated:", datetime.now().strftime("%d %b %Y, %I:%M %p")),
    ]

    for i, (label, value) in enumerate(details):
        row = 4 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = Font(name="Calibri", size=11, color="333333")
        ws.merge_cells(f"B{row}:E{row}")

    # ── Statistics ────────────────────────────────────────────────────────────
    if submissions:
        marks_list = [s.get("final_marks") or s.get("ai_marks") or s.get("marks_obtained") or 0 for s in submissions]
        max_marks = assignment.get("max_marks", 100)
        avg_marks = sum(marks_list) / len(marks_list) if marks_list else 0
        pass_count = sum(1 for m in marks_list if m >= max_marks * 0.4)
        
        stats = [
            ("Total Submissions:", str(len(submissions))),
            ("Average Score:", f"{avg_marks:.1f} / {max_marks}"),
            ("Highest Score:", f"{max(marks_list)} / {max_marks}"),
            ("Lowest Score:", f"{min(marks_list)} / {max_marks}"),
            ("Pass Count:", f"{pass_count} / {len(submissions)}"),
        ]

        ws["G4"] = "STATISTICS"
        ws["G4"].font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        ws["G4"].fill = PatternFill("solid", fgColor=BLUE)
        ws["G4"].alignment = Alignment(horizontal="center")
        ws.merge_cells("G4:I4")

        for i, (label, value) in enumerate(stats):
            row = 5 + i
            ws[f"G{row}"] = label
            ws[f"G{row}"].font = Font(name="Calibri", size=10, bold=True, color=NAVY)
            ws[f"H{row}"] = value
            ws[f"H{row}"].font = Font(name="Calibri", size=10)
            ws[f"H{row}"].fill = PatternFill("solid", fgColor=LIGHT_BG)
            ws.merge_cells(f"H{row}:I{row}")

    # ── Table Header ─────────────────────────────────────────────────────────
    header_row = 12
    headers = ["#", "Roll Number", "Student Name", "Email", "Type",
               "AI Marks", "Final Marks", "Max", "Percentage", "Source", "Feedback", "Review?"]
    col_widths = [4, 20, 22, 28, 12, 8, 10, 6, 12, 12, 50, 10]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[header_row].height = 25

    # ── Student Rows ──────────────────────────────────────────────────────────
    for idx, sub in enumerate(submissions, 1):
        row = header_row + idx
        marks = sub.get("final_marks") or sub.get("ai_marks") or sub.get("marks_obtained") or 0
        max_m = sub.get("max_marks") or assignment.get("max_marks", 100)
        pct = (marks / max_m * 100) if max_m > 0 else 0
        needs_review = bool(sub.get("needs_review"))

        # Color-code rows by grade
        if pct >= 75:
            row_fill = PatternFill("solid", fgColor=GREEN_BG)
            grade_fill = PatternFill("solid", fgColor="2E7D32")
            grade_color = WHITE
        elif pct >= 40:
            row_fill = PatternFill("solid", fgColor=YELLOW_BG)
            grade_fill = PatternFill("solid", fgColor="F57F17")
            grade_color = WHITE
        else:
            row_fill = PatternFill("solid", fgColor=RED_BG)
            grade_fill = PatternFill("solid", fgColor="C62828")
            grade_color = WHITE

        ai_marks    = sub.get("ai_marks") or sub.get("marks_obtained") or 0
        final_marks = sub.get("final_marks") or ai_marks
        source      = "Teacher Override" if sub.get("teacher_marks") is not None else "AI Evaluated"
        feedback    = (sub.get("teacher_feedback") or sub.get("ai_feedback") or sub.get("feedback") or "")[:200]

        row_data = [
            idx,
            sub.get("roll_number", ""),
            sub.get("student_name", ""),
            sub.get("email", ""),
            sub.get("submission_type", "typed").title(),
            ai_marks,
            final_marks,
            max_m,
            f"{pct:.1f}%",
            source,
            feedback,
            "⚠ Yes" if needs_review else "✓ No"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 9))

            # Apply row background
            if col not in [6, 10]:
                cell.fill = row_fill

            # Highlight Final Marks column
            if col == 7:
                cell.fill = grade_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color=grade_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Special: Review column
            if col == 12:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if needs_review:
                    cell.fill = PatternFill("solid", fgColor="FFE0B2")
                    cell.font = Font(name="Calibri", size=10, color="E65100")
                else:
                    cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                    cell.font = Font(name="Calibri", size=10, color="2E7D32")

        ws.row_dimensions[row].height = 45

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Sheet 2: Needs Review ─────────────────────────────────────────────────
    flagged = [s for s in submissions if s.get("needs_review")]
    if flagged:
        ws2 = wb.create_sheet("⚠ Needs Review")
        ws2["A1"] = "SUBMISSIONS FLAGGED FOR MANUAL REVIEW"
        ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
        ws2["A1"].fill = PatternFill("solid", fgColor="C62828")
        ws2.merge_cells("A1:F1")
        ws2.row_dimensions[1].height = 35

        ws2["A2"] = "These submissions could not be fully evaluated automatically. Please review manually."
        ws2["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
        ws2.merge_cells("A2:F2")

        r2_headers = ["Roll Number", "Student Name", "Type", "Marks (AI)", "Feedback", "Submitted At"]
        r2_widths = [20, 25, 12, 12, 60, 20]
        for col, (h, w) in enumerate(zip(r2_headers, r2_widths), 1):
            c = ws2.cell(row=4, column=col, value=h)
            c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="C62828")
            c.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col)].width = w

        for i, sub in enumerate(flagged, 1):
            data = [
                sub.get("roll_number", ""),
                sub.get("student_name", ""),
                sub.get("submission_type", "").title(),
                f"{sub.get('final_marks') or sub.get('ai_marks') or sub.get('marks_obtained') or 0} / {sub.get('max_marks', 100)}",
                (sub.get("teacher_feedback") or sub.get("ai_feedback") or sub.get("feedback") or "")[:300],
                sub.get("submitted_at", "")[:16].replace("T", " ")
            ]
            for col, val in enumerate(data, 1):
                c = ws2.cell(row=4+i, column=col, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor="FFF3E0")
                c.border = make_border()
                c.alignment = Alignment(vertical="center", wrap_text=(col == 5))
            ws2.row_dimensions[4+i].height = 50

    # ── Save file ─────────────────────────────────────────────────────────────
    output_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f"Results_{assignment.get('assignment_id', 'export')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    print(f"📊 Excel exported: {filepath}")
    return filepath


def make_border(color="CCCCCC"):
    """Helper: Create a thin border for table cells"""
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)
